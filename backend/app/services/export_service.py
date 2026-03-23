import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from app.models.asistencia import Asistencia
from app.models.alumno import Alumno
from app.models.grado import Grado
from app.utils.timezone import get_peru_today # Added import

def exportar_asistencias_excel(
    db: Session,
    grado_id: int,
    fecha_inicio: date | None = None,
    fecha_fin: date | None = None,
) -> io.BytesIO:
    """Export attendance records for a grade as an Excel file."""
    # Get grado info
    grado = db.query(Grado).filter(Grado.id == grado_id).first()
    grado_nombre = grado.nombre if grado else "Desconocido"

    # Default date range: current month
    # Modified to use get_peru_today()
    peru_today = get_peru_today()
    if not fecha_inicio:
        fecha_inicio = peru_today.replace(day=1)
    if not fecha_fin:
        fecha_fin = peru_today

    # Query attendance data
    asistencias = (
        db.query(Asistencia)
        .join(Alumno, Asistencia.alumno_dni == Alumno.dni)
        .filter(
            Alumno.grado_id == grado_id,
            Asistencia.fecha >= fecha_inicio,
            Asistencia.fecha <= fecha_fin,
        )
        .order_by(Alumno.apellidos, Alumno.nombres, Asistencia.fecha)
        .all()
    )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Asistencia - {grado_nombre}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Reporte de Asistencia - {grado_nombre}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    # Date range row
    ws.merge_cells("A2:F2")
    date_cell = ws["A2"]
    date_cell.value = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    date_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["DNI", "Apellidos", "Nombres", "Fecha", "Estado", "Hora Registro"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    estado_fills = {
        "asistencia": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "tardanza": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "inasistencia": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "justificacion": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    }

    for row_idx, asist in enumerate(asistencias, 5):
        ws.cell(row=row_idx, column=1, value=asist.alumno_dni).border = thin_border
        ws.cell(row=row_idx, column=2, value=asist.alumno.apellidos).border = thin_border
        ws.cell(row=row_idx, column=3, value=asist.alumno.nombres).border = thin_border
        ws.cell(row=row_idx, column=4, value=asist.fecha.strftime("%d/%m/%Y")).border = thin_border

        estado_cell = ws.cell(row=row_idx, column=5, value=asist.estado.value.capitalize())
        estado_cell.border = thin_border
        estado_cell.fill = estado_fills.get(asist.estado.value, PatternFill())

        hora = asist.hora_registro.strftime("%H:%M:%S") if asist.hora_registro else "-"
        ws.cell(row=row_idx, column=6, value=hora).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, grado_nombre


def generar_excel_general(
    db: Session,
    grado_ids: list[int] | None = None,
    fecha_inicio: date | None = None,
    fecha_fin: date | None = None,
) -> tuple[io.BytesIO, str]:
    """Export attendance records for all assigned grades as a multi-sheet Excel file."""
    
    if grado_ids:
        grados = db.query(Grado).filter(Grado.id.in_(grado_ids)).order_by(Grado.id).all()
    else:
        grados = db.query(Grado).order_by(Grado.id).all()

    peru_today = get_peru_today()
    if not fecha_inicio:
        fecha_inicio = peru_today.replace(day=1)
    if not fecha_fin:
        fecha_fin = peru_today

    wb = Workbook()
    # Remove the default sheet as we will create one for each grade
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    estado_fills = {
        "asistencia": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "tardanza": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "inasistencia": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "justificacion": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
    }

    if not grados:
        # Fallback if no grades assigned
        ws = wb.create_sheet(title="Vacio")
        ws.cell(row=1, column=1, value="No hay grados asignados")
    else:
        for grado in grados:
            # Query attendance data for this grade
            asistencias = (
                db.query(Asistencia)
                .join(Alumno, Asistencia.alumno_dni == Alumno.dni)
                .filter(
                    Alumno.grado_id == grado.id,
                    Asistencia.fecha >= fecha_inicio,
                    Asistencia.fecha <= fecha_fin,
                )
                .order_by(Alumno.apellidos, Alumno.nombres, Asistencia.fecha)
                .all()
            )

            # Create sheet name safely (max 31 chars, invalid chars removed by openpyxl if needed, but names here should be safe)
            ws = wb.create_sheet(title=f"{grado.nombre}")

            # Title row
            ws.merge_cells("A1:F1")
            title_cell = ws["A1"]
            title_cell.value = f"Reporte de Asistencia - {grado.nombre}"
            title_cell.font = Font(bold=True, size=14, color="1F4E79")
            title_cell.alignment = Alignment(horizontal="center")

            # Date range row
            ws.merge_cells("A2:F2")
            date_cell = ws["A2"]
            date_cell.value = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            date_cell.alignment = Alignment(horizontal="center")

            # Headers
            headers = ["DNI", "Apellidos", "Nombres", "Fecha", "Estado", "Hora Registro"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data rows
            for row_idx, asist in enumerate(asistencias, 5):
                ws.cell(row=row_idx, column=1, value=asist.alumno_dni).border = thin_border
                ws.cell(row=row_idx, column=2, value=asist.alumno.apellidos).border = thin_border
                ws.cell(row=row_idx, column=3, value=asist.alumno.nombres).border = thin_border
                ws.cell(row=row_idx, column=4, value=asist.fecha.strftime("%d/%m/%Y")).border = thin_border

                estado_cell = ws.cell(row=row_idx, column=5, value=asist.estado.value.capitalize())
                estado_cell.border = thin_border
                estado_cell.fill = estado_fills.get(asist.estado.value, PatternFill())

                hora = asist.hora_registro.strftime("%H:%M:%S") if asist.hora_registro else "-"
                ws.cell(row=row_idx, column=6, value=hora).border = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 16
            ws.column_dimensions["F"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, "General"
