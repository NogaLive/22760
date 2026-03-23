from fastapi import APIRouter, Depends, HTTPException, status, Response, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from typing import List
import openpyxl
import io
from app.database import get_db
from app.middleware.auth_middleware import get_current_user
from app.models.docente import Docente
from app.models.alumno import Alumno
from app.models.grado import Grado
from app.schemas.alumno import (
    CrearAlumnoRequest,
    ActualizarAlumnoRequest,
    AlumnoOut,
    AlumnoListResponse,
)
from app.services.qr_service import generate_qr_token, generate_qr_image
from app.redis_client import invalidate_cache

router = APIRouter(prefix="/api/alumnos", tags=["Alumnos"])


@router.get("/grado/{grado_id}", response_model=AlumnoListResponse)
def listar_alumnos_por_grado(
    grado_id: int,
    db: Session = Depends(get_db),
    current_user: Docente = Depends(get_current_user),
):
    """Listar todos los alumnos de un grado."""
    from app.redis_client import get_cached, set_cached
    cache_key = f"alumnos:grado:{grado_id}"
    cached = get_cached(cache_key)
    if cached:
        return AlumnoListResponse(**cached)

    grado = db.query(Grado).filter(Grado.id == grado_id).first()
    if not grado:
        raise HTTPException(status_code=404, detail="Grado no encontrado")

    alumnos = (
        db.query(Alumno)
        .filter(Alumno.grado_id == grado_id, Alumno.activo == True)
        .order_by(Alumno.apellidos, Alumno.nombres)
        .all()
    )

    result = [
        AlumnoOut(
            dni=a.dni,
            nombres=a.nombres,
            apellidos=a.apellidos,
            grado_id=a.grado_id,
            grado_nombre=grado.nombre,
            codigo_qr=a.codigo_qr,
            activo=a.activo,
        )
        for a in alumnos
    ]

    response = AlumnoListResponse(alumnos=result, total=len(result))
    set_cached(cache_key, response.model_dump(), ttl=300) # cache for 5 minutes
    return response


@router.post("/importar/{grado_id}")
def importar_alumnos(
    grado_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Docente = Depends(get_current_user),
):
    """Importar alumnos masivamente desde un archivo Excel."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")

    grado = db.query(Grado).filter(Grado.id == grado_id).first()
    if not grado:
        raise HTTPException(status_code=404, detail="Grado no encontrado")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        alumnos_creados = 0
        errores = []

        # Saltamos la cabecera (fila 1)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row): continue
            
            dni, nombres, apellidos = row[0:3]
            
            if not dni or not nombres or not apellidos:
                errores.append(f"Fila {row_idx}: Datos incompletos")
                continue

            # Validar DNI existente
            dni_str = str(dni).strip()
            existing = db.query(Alumno).filter(Alumno.dni == dni_str).first()
            if existing:
                errores.append(f"Fila {row_idx}: DNI {dni_str} ya existe")
                continue

            # Crear alumno
            qr_token = generate_qr_token()
            nuevo_alumno = Alumno(
                dni=dni_str,
                nombres=str(nombres).strip(),
                apellidos=str(apellidos).strip(),
                grado_id=grado_id,
                codigo_qr=qr_token,
                activo=True
            )
            db.add(nuevo_alumno)
            alumnos_creados += 1

        db.commit()
        
        # Invalidar cache
        invalidate_cache(f"alumnos:grado:{grado_id}")
        invalidate_cache("dashboard:*")

        return {
            "message": f"Se han importado {alumnos_creados} alumnos exitosamente.",
            "errores": errores
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")

@router.get("/modelo-excel", response_class=Response)
def descargar_modelo_excel():
    """Generar y descargar una plantilla Excel .xlsx para la importación masiva de alumnos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    # Cabeceras requeridas
    cabeceras = ["DNI", "Nombres", "Apellidos"]
    ws.append(cabeceras)

    # Estilos básicos
    for col in range(1, 4):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Datos de ejemplo
    ejemplos = [
        ["78901234", "Juan Carlos", "Pérez López"],
        ["65432109", "María", "Gómez"]
    ]
    for ejemplo in ejemplos:
        ws.append(ejemplo)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="Plantilla_Alumnos.xlsx"'
    }
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("", response_model=AlumnoOut, status_code=status.HTTP_201_CREATED)
def crear_alumno(
    request: CrearAlumnoRequest,
    db: Session = Depends(get_db),
    current_user: Docente = Depends(get_current_user),
):
    """Crear un nuevo alumno y generar su código QR automáticamente."""
    # Check if DNI already exists
    existing = db.query(Alumno).filter(Alumno.dni == request.dni).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ya existe un alumno con DNI {request.dni}",
        )

    # Verify grado exists
    grado = db.query(Grado).filter(Grado.id == request.grado_id).first()
    if not grado:
        raise HTTPException(status_code=404, detail="Grado no encontrado")

    # Generate unique QR token
    qr_token = generate_qr_token()

    alumno = Alumno(
        dni=request.dni,
        nombres=request.nombres,
        apellidos=request.apellidos,
        grado_id=request.grado_id,
        codigo_qr=qr_token,
        activo=True,
    )

    db.add(alumno)
    db.commit()
    db.refresh(alumno)

    return AlumnoOut(
        dni=alumno.dni,
        nombres=alumno.nombres,
        apellidos=alumno.apellidos,
        grado_id=alumno.grado_id,
        grado_nombre=grado.nombre,
        codigo_qr=alumno.codigo_qr,
        activo=alumno.activo,
    )


@router.put("/{dni}", response_model=AlumnoOut)
def actualizar_alumno(
    dni: int,
    request: ActualizarAlumnoRequest,
    db: Session = Depends(get_db),
    current_user: Docente = Depends(get_current_user),
):
    """Actualizar datos de un alumno (DNI, nombres, apellidos)."""
    alumno = db.query(Alumno).filter(Alumno.dni == dni).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    # If changing DNI, check uniqueness
    if request.dni and request.dni != alumno.dni:
        existing = db.query(Alumno).filter(Alumno.dni == request.dni).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Ya existe un alumno con DNI {request.dni}",
            )
        alumno.dni = request.dni

    if request.nombres:
        alumno.nombres = request.nombres
    if request.apellidos:
        alumno.apellidos = request.apellidos

    db.commit()
    db.refresh(alumno)

    grado = db.query(Grado).filter(Grado.id == alumno.grado_id).first()

    return AlumnoOut(
        dni=alumno.dni,
        nombres=alumno.nombres,
        apellidos=alumno.apellidos,
        grado_id=alumno.grado_id,
        grado_nombre=grado.nombre if grado else "",
        codigo_qr=alumno.codigo_qr,
        activo=alumno.activo,
    )


@router.delete("/{dni}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_alumno(
    dni: int,
    db: Session = Depends(get_db),
    current_user: Docente = Depends(get_current_user),
):
    """Eliminar (desactivar) un alumno del grado."""
    alumno = db.query(Alumno).filter(Alumno.dni == dni).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    alumno.activo = False
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/{dni}/qr")
def obtener_qr_alumno(
    dni: int,
    db: Session = Depends(get_db),
    current_user: Docente = Depends(get_current_user),
):
    """Obtener imagen QR de un alumno como PNG."""
    alumno = db.query(Alumno).filter(Alumno.dni == dni, Alumno.activo == True).first()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    qr_bytes = generate_qr_image(alumno.codigo_qr)

    return Response(
        content=qr_bytes,
        media_type="image/png",
        headers={
            "Content-Disposition": f"inline; filename=qr_{alumno.dni}.png",
        },
    )
